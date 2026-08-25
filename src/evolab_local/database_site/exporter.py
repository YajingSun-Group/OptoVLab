from __future__ import annotations

import copy
import gzip
import hashlib
import json
import re
import sqlite3
from collections import defaultdict
from collections.abc import Iterable, Mapping
from dataclasses import asdict, dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


PATH_PART_RE = re.compile(r"^([A-Za-z_][A-Za-z0-9_]*)(?:\[(\d+)\])?$")


@dataclass(frozen=True)
class DatasetManifest:
    key: str
    label: str
    record_count: int
    paper_count: int
    file: str
    compressed_bytes: int
    sha256: str
    source: str
    source_url: str | None = None
    license: str | None = None
    quality_counts: dict[str, int] | None = None


@dataclass(frozen=True)
class ExportManifest:
    schema_version: str
    generated_at: str
    datasets: list[DatasetManifest]


def export_static_database_data(
    *,
    oled_database: Path,
    ofet_workbook: Path,
    opv_json: Path,
    output_directory: Path,
) -> ExportManifest:
    output_directory.mkdir(parents=True, exist_ok=True)
    generated_at = datetime.now(UTC).replace(microsecond=0).isoformat()

    oled_records, oled_metadata = export_oled_records(oled_database)
    ofet_records = export_ofet_records(ofet_workbook)
    opv_records = read_json_array(opv_json)

    manifests = [
        _write_dataset(
            key="oled",
            label="Organic Light-Emitting Diodes",
            records=oled_records,
            output_directory=output_directory,
            paper_count=int(oled_metadata["paper_count"]),
            source="Reviewed OptoVLab mining-platform SQLite export",
            quality_counts=dict(oled_metadata["quality_counts"]),
        ),
        _write_dataset(
            key="ofet",
            label="Organic Field-Effect Transistors",
            records=ofet_records,
            output_directory=output_directory,
            paper_count=len({record["doi"] for record in ofet_records if record.get("doi")}),
            source="User-supplied OFET workbook export",
        ),
        _write_dataset(
            key="opv",
            label="Organic Photovoltaics",
            records=opv_records,
            output_directory=output_directory,
            paper_count=len(
                {
                    record.get("doi_norm") or record.get("doi")
                    for record in opv_records
                    if record.get("doi_norm") or record.get("doi")
                }
            ),
            source="OPV-DB publication release 1.0.0",
            source_url="https://github.com/sunyrain/OPV2D",
            license="CC BY 4.0",
            quality_counts={
                "strict_performance": sum(
                    bool(record.get("strict_performance_benchmark"))
                    for record in opv_records
                ),
                "strict_molecular": sum(
                    bool(record.get("strict_molecular_benchmark"))
                    for record in opv_records
                ),
            },
        ),
    ]
    manifest = ExportManifest(
        schema_version="1.0",
        generated_at=generated_at,
        datasets=manifests,
    )
    _write_json(output_directory / "catalog.json", _manifest_payload(manifest))
    return manifest


def export_oled_records(database_path: Path) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    connection = sqlite3.connect(f"file:{database_path.as_posix()}?mode=ro", uri=True)
    connection.row_factory = sqlite3.Row
    try:
        papers = {
            row["paper_id"]: dict(row)
            for row in connection.execute(
                """
                SELECT paper_id, doi, title, journal, publisher, year, review_status,
                       review_reason, updated_at
                FROM papers
                WHERE review_status = 'confirmed'
                  AND review_reason IS NULL
                """
            )
        }
        final_records = {
            row["paper_id"]: dict(row)
            for row in connection.execute(
                """
                SELECT paper_id, candidate_run_id, final_json, confirmed_at
                FROM candidate_final_records
                WHERE status = 'confirmed'
                ORDER BY confirmed_at DESC
                """
            )
            if row["paper_id"] in papers
        }
        latest_runs = _latest_completed_runs(connection, set(papers))
        reviewed_values = _reviewed_values_by_run(connection, set(latest_runs.values()))
        material_structures = _material_structures_by_run(
            connection,
            set(latest_runs.values()) | {
                row["candidate_run_id"] for row in final_records.values()
            },
        )

        records: list[dict[str, Any]] = []
        included_papers: set[str] = set()
        quality_counts: dict[str, int] = defaultdict(int)
        for paper_id, paper in papers.items():
            final_record = final_records.get(paper_id)
            if final_record:
                payload = _json_object(final_record["final_json"])
                candidate_run_id = final_record["candidate_run_id"]
                quality_tier = "human_finalized"
                finalized_at = final_record["confirmed_at"]
            else:
                candidate_run_id = latest_runs.get(paper_id)
                if not candidate_run_id:
                    continue
                run = connection.execute(
                    """
                    SELECT mining_result_json, completed_at
                    FROM candidate_ingestion_runs
                    WHERE candidate_run_id = ?
                    """,
                    (candidate_run_id,),
                ).fetchone()
                if not run:
                    continue
                payload = _json_object(run["mining_result_json"])
                _apply_reviewed_values(payload, reviewed_values.get(candidate_run_id, []))
                quality_tier = "auto_reviewed"
                finalized_at = None

            devices = payload.get("devices")
            if not isinstance(devices, list) or not devices:
                continue
            paper_materials = _materials_by_id(payload.get("materials"))
            run_materials = material_structures.get(candidate_run_id, {})
            for material_id, material in paper_materials.items():
                structure = run_materials.get(material_id)
                if structure:
                    material.update(structure)

            for device_index, device in enumerate(devices):
                if not isinstance(device, Mapping):
                    continue
                record = _flatten_oled_device(
                    paper=paper,
                    device=dict(device),
                    paper_materials=paper_materials,
                    candidate_run_id=candidate_run_id,
                    device_index=device_index,
                    quality_tier=quality_tier,
                    finalized_at=finalized_at,
                )
                records.append(record)
                quality_counts[quality_tier] += 1
                included_papers.add(paper_id)

        records.sort(
            key=lambda record: (
                -(record.get("year") or 0),
                str(record.get("doi") or ""),
                int(record.get("device_index") or 0),
            )
        )
        return records, {
            "paper_count": len(included_papers),
            "quality_counts": dict(quality_counts),
        }
    finally:
        connection.close()


def export_ofet_records(workbook_path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = worksheet.iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(rows)]
    records: list[dict[str, Any]] = []
    for index, values in enumerate(rows, start=1):
        source = dict(zip(headers, values, strict=False))
        if not any(value is not None and str(value).strip() for value in values):
            continue
        record = {
            "id": index,
            "smiles": _clean_cell(source.get("smiles")),
            "semiconductor": _clean_cell(source.get("Organic Semiconductor Layer")),
            "fabrication_method": _clean_cell(source.get("Fabrication Method")),
            "organic_layer_thickness": _clean_cell(source.get("Organic Layer Thickness")),
            "fabrication_details": _clean_cell(
                source.get("OFET Device Fabrication Details")
            ),
            "source_electrode": _clean_cell(source.get("Source Electrodes")),
            "drain_electrode": _clean_cell(source.get("Drain Electrodes")),
            "dielectric_layer": _clean_cell(source.get("Dielectric Layer")),
            "dielectric_thickness": _clean_cell(
                source.get("Dielectric Layer Thickness")
            ),
            "gate_electrode": _clean_cell(source.get("Gate Electrode")),
            "device_geometry": _clean_cell(source.get("Device Geometries")),
            "test_atmosphere": _clean_cell(source.get("Test Atmosphere")),
            "on_off_ratio": _clean_cell(
                source.get("On-to-Off Current Ratios (I_on/I_off)")
            ),
            "threshold_voltage": _clean_cell(source.get("Threshold Voltage")),
            "conduction_type": _clean_cell(source.get("Conduction Type")),
            "mobility": _clean_cell(source.get("Mobility")),
            "highest_mobility": _number_or_none(source.get("Highest Mobility")),
            "pdf": _clean_cell(source.get("pdf")),
            "directory": _clean_cell(source.get("dir")),
            "doi": _normalize_doi(source.get("doi")),
            "mobility_bucket": _clean_cell(source.get("Mobility Cut")),
            "year": _integer_or_none(source.get("publication_year")),
            "major_category": _clean_cell(source.get("major_category")),
            "sub_category": _clean_cell(source.get("sub_category")),
            "fabrication_category": _clean_cell(source.get("Fabrication Category")),
            "source_electrode_category": _clean_cell(
                source.get("Source Electrodes Category")
            ),
            "drain_electrode_category": _clean_cell(
                source.get("Drain Electrodes Category")
            ),
            "gate_electrode_category": _clean_cell(
                source.get("Gate Electrode Category")
            ),
            "dielectric_category": _clean_cell(
                source.get("Dielectric Layer Category")
            ),
            "conduction_category": _clean_cell(
                source.get("Conduction Type Category")
            ),
            "geometry_category": _clean_cell(
                source.get("Device Geometries Category")
            ),
            "publisher": _clean_cell(source.get("publisher")),
            "journal": _clean_cell(source.get("journal")),
            "atmosphere_category": _clean_cell(
                source.get("Test Atmosphere Category")
            ),
        }
        record["has_smiles"] = bool(record["smiles"])
        records.append(record)
    workbook.close()
    return records


def read_json_array(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8") as handle:
        payload = json.load(handle)
    if not isinstance(payload, list) or not all(isinstance(row, dict) for row in payload):
        raise ValueError(f"Expected a JSON array of objects: {path}")
    return payload


def _latest_completed_runs(
    connection: sqlite3.Connection,
    paper_ids: set[str],
) -> dict[str, str]:
    if not paper_ids:
        return {}
    latest: dict[str, str] = {}
    for row in connection.execute(
        """
        SELECT paper_id, candidate_run_id
        FROM candidate_ingestion_runs
        WHERE status = 'completed'
        ORDER BY paper_id, COALESCE(completed_at, created_at) DESC, created_at DESC
        """
    ):
        paper_id = row["paper_id"]
        if paper_id in paper_ids and paper_id not in latest:
            latest[paper_id] = row["candidate_run_id"]
    return latest


def _reviewed_values_by_run(
    connection: sqlite3.Connection,
    candidate_run_ids: set[str],
) -> dict[str, list[dict[str, Any]]]:
    reviewed: dict[str, list[dict[str, Any]]] = defaultdict(list)
    if not candidate_run_ids:
        return reviewed
    for row in connection.execute(
        """
        SELECT candidate_run_id, concrete_path, value_json, reviewed_value_json, status
        FROM candidate_values
        WHERE reviewed_value_json IS NOT NULL OR status != 'pending'
        """
    ):
        if row["candidate_run_id"] in candidate_run_ids:
            reviewed[row["candidate_run_id"]].append(dict(row))
    return reviewed


def _material_structures_by_run(
    connection: sqlite3.Connection,
    candidate_run_ids: set[str],
) -> dict[str, dict[str, dict[str, Any]]]:
    structures: dict[str, dict[str, dict[str, Any]]] = defaultdict(dict)
    if not candidate_run_ids:
        return structures
    for row in connection.execute(
        """
        SELECT l.candidate_run_id, l.paper_material_id, l.match_method,
               l.match_confidence, l.match_status, g.global_material_id,
               g.canonical_name, g.material_class, g.representation_type,
               g.canonical_smiles, g.isomeric_smiles, g.inchi, g.inchi_key,
               g.formula, g.molecular_weight, g.source, g.confidence,
               g.review_status
        FROM paper_material_links AS l
        LEFT JOIN materials_global AS g
          ON g.global_material_id = l.global_material_id
        WHERE l.global_material_id IS NOT NULL
        """
    ):
        run_id = row["candidate_run_id"]
        if run_id not in candidate_run_ids:
            continue
        structures[run_id][row["paper_material_id"]] = {
            "global_material_id": row["global_material_id"],
            "canonical_name": row["canonical_name"],
            "material_class": row["material_class"],
            "representation_type": row["representation_type"],
            "canonical_smiles": row["canonical_smiles"],
            "isomeric_smiles": row["isomeric_smiles"],
            "inchi": row["inchi"],
            "inchi_key": row["inchi_key"],
            "formula": row["formula"],
            "molecular_weight": row["molecular_weight"],
            "structure_source": row["source"],
            "structure_confidence": row["confidence"],
            "structure_review_status": row["review_status"],
            "match_method": row["match_method"],
            "match_confidence": row["match_confidence"],
            "match_status": row["match_status"],
        }
    return structures


def _apply_reviewed_values(
    payload: dict[str, Any],
    reviewed_values: Iterable[Mapping[str, Any]],
) -> None:
    for value in reviewed_values:
        if value.get("status") == "rejected":
            reviewed_value = None
        elif value.get("reviewed_value_json") is not None:
            reviewed_value = json.loads(str(value["reviewed_value_json"]))
        else:
            reviewed_value = json.loads(str(value["value_json"]))
        _set_by_path(payload, str(value["concrete_path"]), reviewed_value)


def _set_by_path(payload: dict[str, Any], path: str, value: Any) -> None:
    parts = path.split(".")
    current: Any = payload
    for offset, raw_part in enumerate(parts):
        match = PATH_PART_RE.fullmatch(raw_part)
        if not match:
            return
        key, index_text = match.groups()
        is_last = offset == len(parts) - 1
        if not isinstance(current, dict) or key not in current:
            return
        if index_text is None:
            if is_last:
                current[key] = value
                return
            current = current[key]
            continue
        collection = current.get(key)
        index = int(index_text)
        if not isinstance(collection, list) or index >= len(collection):
            return
        if is_last:
            collection[index] = value
            return
        current = collection[index]


def _materials_by_id(value: Any) -> dict[str, dict[str, Any]]:
    materials: dict[str, dict[str, Any]] = {}
    if not isinstance(value, list):
        return materials
    for item in value:
        if not isinstance(item, Mapping):
            continue
        material_id = item.get("paper_material_id")
        if isinstance(material_id, str) and material_id:
            materials[material_id] = copy.deepcopy(dict(item))
    return materials


def _flatten_oled_device(
    *,
    paper: Mapping[str, Any],
    device: dict[str, Any],
    paper_materials: Mapping[str, dict[str, Any]],
    candidate_run_id: str,
    device_index: int,
    quality_tier: str,
    finalized_at: str | None,
) -> dict[str, Any]:
    layers = [
        copy.deepcopy(dict(layer))
        for layer in device.get("layers", [])
        if isinstance(layer, Mapping)
    ]
    performance = [
        copy.deepcopy(dict(item))
        for item in device.get("performance", [])
        if isinstance(item, Mapping)
    ]
    used_material_ids = _device_material_ids(device)
    materials = [
        copy.deepcopy(paper_materials[material_id])
        for material_id in sorted(used_material_ids)
        if material_id in paper_materials
    ]
    final_emitter = device.get("final_emitter")
    final_emitter = dict(final_emitter) if isinstance(final_emitter, Mapping) else {}
    final_emitter_id = final_emitter.get("paper_material_id")
    final_emitter_material = (
        paper_materials.get(final_emitter_id, {})
        if isinstance(final_emitter_id, str)
        else {}
    )
    doi = _normalize_doi(paper.get("doi"))
    return {
        "id": f"{paper['paper_id']}::D{device_index + 1}",
        "paper_id": paper["paper_id"],
        "candidate_run_id": candidate_run_id,
        "device_index": device_index,
        "doi": doi,
        "title": paper.get("title"),
        "journal": paper.get("journal"),
        "publisher": paper.get("publisher"),
        "year": paper.get("year"),
        "quality_tier": quality_tier,
        "finalized_at": finalized_at,
        "device_label": device.get("device_label") or f"Device {device_index + 1}",
        "device_type": device.get("device_type"),
        "control_or_target": device.get("control_or_target"),
        "architecture": device.get("architecture_text"),
        "emission_color": device.get("emission_color"),
        "emission_mechanism": _string_list(device.get("emission_mechanism")),
        "fabrication_method": _nested_value(device, "fabrication", "method"),
        "is_tandem": bool(device.get("is_tandem")),
        "is_white_oled": bool(device.get("is_white_oled")),
        "layer_count": len(layers),
        "material_count": len(used_material_ids),
        "final_emitter": final_emitter.get("mention"),
        "final_emitter_class": final_emitter.get("material_class"),
        "final_emitter_smiles": final_emitter_material.get("canonical_smiles")
        or final_emitter_material.get("isomeric_smiles"),
        "eqe_max": _performance_value(performance, "EQE", "max"),
        "ce_max": _performance_value(performance, "CE", "max"),
        "pe_max": _performance_value(performance, "PE", "max"),
        "luminance_max": _performance_value(performance, "luminance", "max"),
        "turn_on_voltage": _turn_on_voltage(performance),
        "el_peak": _performance_match(performance, ("EL peak", "peak wavelength")),
        "fwhm": _performance_match(performance, ("FWHM",)),
        "lifetime": _performance_match(performance, ("LT50", "LT80", "LT95", "lifetime")),
        "layers": layers,
        "performance": performance,
        "materials": materials,
    }


def _device_material_ids(device: Mapping[str, Any]) -> set[str]:
    material_ids: set[str] = set()
    final_emitter = device.get("final_emitter")
    if isinstance(final_emitter, Mapping):
        material_id = final_emitter.get("paper_material_id")
        if isinstance(material_id, str) and material_id:
            material_ids.add(material_id)
    layers = device.get("layers")
    if not isinstance(layers, list):
        return material_ids
    for layer in layers:
        if not isinstance(layer, Mapping):
            continue
        components = layer.get("components")
        if not isinstance(components, list):
            continue
        for component in components:
            if not isinstance(component, Mapping):
                continue
            material_id = component.get("paper_material_id")
            if isinstance(material_id, str) and material_id:
                material_ids.add(material_id)
    return material_ids


def _performance_value(
    performance: list[dict[str, Any]],
    metric_family: str,
    statistic: str,
) -> float | None:
    for item in performance:
        if (
            str(item.get("metric_family") or "").casefold() == metric_family.casefold()
            and str(item.get("statistic") or "").casefold() == statistic.casefold()
        ):
            return _number_or_none(item.get("normalized_value"))
    return None


def _performance_match(
    performance: list[dict[str, Any]],
    names: tuple[str, ...],
) -> float | None:
    needles = tuple(name.casefold() for name in names)
    for item in performance:
        label = " ".join(
            str(item.get(key) or "") for key in ("metric_name", "metric_family")
        ).casefold()
        if any(needle in label for needle in needles):
            return _number_or_none(item.get("normalized_value"))
    return None


def _turn_on_voltage(performance: list[dict[str, Any]]) -> float | None:
    for item in performance:
        label = str(item.get("metric_name") or "").casefold()
        if "turn-on" in label or "turn_on" in label or "turn on" in label:
            return _number_or_none(item.get("normalized_value"))
    return None


def _nested_value(value: Mapping[str, Any], key: str, nested_key: str) -> Any:
    nested = value.get(key)
    return nested.get(nested_key) if isinstance(nested, Mapping) else None


def _string_list(value: Any) -> list[str]:
    if isinstance(value, list):
        return [str(item) for item in value if item is not None and str(item).strip()]
    if value is None or not str(value).strip():
        return []
    rendered = str(value).strip()
    if rendered.startswith("[") and rendered.endswith("]"):
        try:
            parsed = json.loads(rendered)
        except json.JSONDecodeError:
            parsed = None
        if isinstance(parsed, list):
            return [
                str(item) for item in parsed if item is not None and str(item).strip()
            ]
    return [rendered]


def _normalize_doi(value: Any) -> str | None:
    text = _clean_cell(value)
    if not text:
        return None
    lowered = text.casefold()
    for prefix in ("https://doi.org/", "http://doi.org/", "doi:"):
        if lowered.startswith(prefix):
            return text[len(prefix) :].strip()
    return text


def _clean_cell(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (int, float, bool)):
        return value
    text = str(value).strip()
    return text or None


def _number_or_none(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number


def _integer_or_none(value: Any) -> int | None:
    number = _number_or_none(value)
    return int(number) if number is not None else None


def _json_object(value: str) -> dict[str, Any]:
    payload = json.loads(value)
    if not isinstance(payload, dict):
        raise ValueError("Expected JSON object")
    return payload


def _write_dataset(
    *,
    key: str,
    label: str,
    records: list[dict[str, Any]],
    output_directory: Path,
    paper_count: int,
    source: str,
    source_url: str | None = None,
    license: str | None = None,
    quality_counts: dict[str, int] | None = None,
) -> DatasetManifest:
    output_path = output_directory / f"{key}.json.gz"
    content = json.dumps(
        records,
        ensure_ascii=False,
        separators=(",", ":"),
        allow_nan=False,
    ).encode("utf-8")
    with output_path.open("wb") as raw_handle:
        with gzip.GzipFile(
            filename="",
            mode="wb",
            compresslevel=9,
            fileobj=raw_handle,
            mtime=0,
        ) as handle:
            handle.write(content)
    compressed = output_path.read_bytes()
    return DatasetManifest(
        key=key,
        label=label,
        record_count=len(records),
        paper_count=paper_count,
        file=f"data/{output_path.name}",
        compressed_bytes=len(compressed),
        sha256=hashlib.sha256(compressed).hexdigest(),
        source=source,
        source_url=source_url,
        license=license,
        quality_counts=quality_counts,
    )


def _write_json(path: Path, payload: Any) -> None:
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )


def _manifest_payload(manifest: ExportManifest) -> dict[str, Any]:
    return {
        "schema_version": manifest.schema_version,
        "generated_at": manifest.generated_at,
        "datasets": [asdict(dataset) for dataset in manifest.datasets],
    }
