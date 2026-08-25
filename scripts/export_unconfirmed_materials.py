from __future__ import annotations

import argparse
from collections import Counter, defaultdict
from datetime import datetime
import json
from pathlib import Path
import re
import sqlite3
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

from evolab_local.mining_platform.core.config import load_config
from evolab_local.mining_platform.material_resolution_service import MaterialResolutionService


COMPLETE_MATERIAL_STATUSES = {"Ready", "Identity only", "Structure not needed"}
EXCLUDED_OCSR_STATUSES = {"OCSR candidate ready", "OCSR correction needed"}
ILLEGAL_EXCEL_XML_CHARACTERS = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Export unresolved materials from active, unconfirmed review-batch papers. "
            "Active DECIMER OCSR candidates can optionally be included in the main sheet."
        )
    )
    parser.add_argument(
        "--config",
        type=Path,
        default=Path("config/mining_platform/mining_platform.yaml"),
    )
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument(
        "--include-ocsr",
        action="store_true",
        help="Include OCSR candidate-ready and correction-needed materials in the main sheet.",
    )
    parser.add_argument("--start-batch", type=int, default=1)
    parser.add_argument("--end-batch", type=int)
    args = parser.parse_args()
    if args.start_batch < 1:
        parser.error("--start-batch must be >= 1")
    if args.end_batch is not None and args.end_batch < args.start_batch:
        parser.error("--end-batch must be >= --start-batch")

    config = load_config(args.config)
    database_path = config.paths.sqlite_path
    batch_size = config.batch_worker.review_batch_size
    generated_at = datetime.now(ZoneInfo("Asia/Shanghai")).isoformat(timespec="seconds")

    with sqlite3.connect(database_path) as conn:
        conn.row_factory = sqlite3.Row
        jobs = _ordered_batch_jobs(conn)
        visual_by_paper = _load_visual_stats(conn)
        bindings_by_material = _load_binding_stats(conn)

    resolution = MaterialResolutionService(config)
    rows: list[dict[str, Any]] = []
    batch_stats: dict[int, Counter[str]] = defaultdict(Counter)
    status_counts: Counter[str] = Counter()
    counters: Counter[str] = Counter()

    active_jobs: list[tuple[int, sqlite3.Row]] = []
    for index, job in enumerate(jobs):
        batch_number = index // batch_size + 1
        if batch_number < args.start_batch or (
            args.end_batch is not None and batch_number > args.end_batch
        ):
            continue
        is_confirmed = job["review_status"] == "confirmed" or int(job["final_count"]) > 0
        if is_confirmed:
            counters["confirmed_papers_excluded"] += 1
            continue
        if job["review_status"] == "excluded":
            counters["excluded_papers_excluded"] += 1
            continue
        active_jobs.append((index, job))
        batch_stats[batch_number]["unconfirmed_papers"] += 1

    print(
        f"Active unconfirmed papers: {len(active_jobs)}; "
        f"review batches: {len(jobs) // batch_size + bool(len(jobs) % batch_size)}",
        flush=True,
    )

    for processed, (index, job) in enumerate(active_jobs, start=1):
        batch_number = index // batch_size + 1
        batch_position = index % batch_size + 1
        bundle = resolution.get_material_structure_bundle(str(job["paper_id"]))
        if bundle is None or not bundle.candidate_run_id or not bundle.materials:
            counters["papers_without_completed_material_run"] += 1
            batch_stats[batch_number]["papers_without_materials"] += 1
            continue
        counters["papers_with_materials"] += 1
        batch_stats[batch_number]["papers_with_materials"] += 1

        links = {item.paper_material_id: item for item in bundle.links}
        tasks = {item.paper_material_id: item for item in bundle.tasks}
        global_materials = {item.global_material_id: item for item in bundle.global_materials}
        candidates_by_material: dict[str, list[Any]] = defaultdict(list)
        for candidate in bundle.structure_candidates:
            candidates_by_material[candidate.paper_material_id].append(candidate)
        judgments_by_candidate: dict[str, Any] = {}
        for judgment in bundle.identity_judgments:
            if judgment.status != "completed":
                continue
            current = judgments_by_candidate.get(judgment.structure_candidate_id)
            if current is None or (judgment.updated_at, judgment.created_at) > (
                current.updated_at,
                current.created_at,
            ):
                judgments_by_candidate[judgment.structure_candidate_id] = judgment
        evidence_runs_by_material: dict[str, list[Any]] = defaultdict(list)
        for evidence_run in bundle.identity_evidence_runs:
            evidence_runs_by_material[evidence_run.paper_material_id].append(evidence_run)
        evidence_items_by_material: dict[str, list[Any]] = defaultdict(list)
        for evidence_item in bundle.identity_evidence_items:
            evidence_items_by_material[evidence_item.paper_material_id].append(evidence_item)
        suggestion_counts = Counter(
            item.paper_material_id for item in bundle.material_name_suggestions
        )

        visual = visual_by_paper.get(str(job["paper_id"]), _empty_visual_stats())
        for material in bundle.materials:
            material_id = material.paper_material_id
            link = links.get(material_id)
            task = tasks.get(material_id)
            linked_global = (
                global_materials.get(link.global_material_id)
                if link and link.global_material_id
                else None
            )
            candidates = candidates_by_material.get(material_id, [])
            evidence_runs = evidence_runs_by_material.get(material_id, [])
            evidence_items = evidence_items_by_material.get(material_id, [])
            binding_stats = bindings_by_material.get(
                (str(job["paper_id"]), material_id),
                {"confirmed": 0, "pending": 0},
            )
            status = _material_status(
                material=material,
                link=link,
                linked_global=linked_global,
                task=task,
                candidates=candidates,
                judgments_by_candidate=judgments_by_candidate,
                evidence_runs=evidence_runs,
                evidence_items=evidence_items,
                binding_stats=binding_stats,
                visual=visual,
                name_suggestion_count=suggestion_counts[material_id],
            )
            counters["device_used_materials"] += 1
            batch_stats[batch_number]["device_used_materials"] += 1
            if status["label"] in COMPLETE_MATERIAL_STATUSES:
                counters["completed_materials_excluded"] += 1
                batch_stats[batch_number]["completed_materials_excluded"] += 1
                continue
            if status["label"] in EXCLUDED_OCSR_STATUSES and not args.include_ocsr:
                counters["ocsr_candidate_materials_excluded"] += 1
                batch_stats[batch_number]["ocsr_candidate_materials_excluded"] += 1
                continue

            row = _material_row(
                job=job,
                batch_number=batch_number,
                batch_position=batch_position,
                candidate_run_id=bundle.candidate_run_id,
                material=material,
                status=status,
                link=link,
                task=task,
                candidates=candidates,
                judgments_by_candidate=judgments_by_candidate,
                evidence_runs=evidence_runs,
                evidence_items=evidence_items,
                binding_stats=binding_stats,
                visual=visual,
                generated_at=generated_at,
            )
            rows.append(row)
            status_counts[status["label"]] += 1
            batch_stats[batch_number]["exported_materials"] += 1

        if processed % 50 == 0 or processed == len(active_jobs):
            print(
                f"Processed {processed}/{len(active_jobs)} papers; "
                f"export rows={len(rows)}; "
                f"OCSR excluded={counters['ocsr_candidate_materials_excluded']}",
                flush=True,
            )

    rows.sort(
        key=lambda item: (
            item["Batch Number"],
            item["Batch Position"],
            item["Material ID"],
        )
    )
    args.output.parent.mkdir(parents=True, exist_ok=True)
    summary = {
        "generated_at": generated_at,
        "database": str(database_path.resolve()),
        "output": str(args.output.resolve()),
        "criteria": {
            "batch_range": {
                "start": args.start_batch,
                "end": args.end_batch,
            },
            "paper": (
                "Review-batch paper is not excluded and is not confirmed by either "
                "papers.review_status or a confirmed candidate_final_record."
            ),
            "material": (
                "Device-used material is not Ready, Identity only, or Structure not needed."
            ),
            "excluded_ocsr": ([] if args.include_ocsr else sorted(EXCLUDED_OCSR_STATUSES)),
        },
        "counters": dict(sorted(counters.items())),
        "exported_row_count": len(rows),
        "status_counts": dict(sorted(status_counts.items())),
    }
    _write_workbook(
        args.output,
        rows=rows,
        summary=summary,
        batch_stats=batch_stats,
    )
    summary_path = args.output.with_suffix(".summary.json")
    summary_path.write_text(
        json.dumps(summary, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(f"Excel: {args.output.resolve()}", flush=True)
    print(f"Summary: {summary_path.resolve()}", flush=True)
    print(json.dumps(summary, ensure_ascii=False, indent=2), flush=True)


def _ordered_batch_jobs(conn: sqlite3.Connection) -> list[sqlite3.Row]:
    return conn.execute(
        """
        WITH final_counts AS (
          SELECT paper_id, COUNT(*) AS final_count
          FROM candidate_final_records
          WHERE status = 'confirmed'
          GROUP BY paper_id
        )
        SELECT
          b.job_id, b.paper_id, b.doi AS job_doi, b.status AS job_status,
          b.current_stage AS job_current_stage, b.created_at AS job_created_at,
          p.doi, p.title, p.journal, p.publisher, p.year,
          p.review_status, p.review_reason, p.parse_status, p.mining_status,
          COALESCE(f.final_count, 0) AS final_count
        FROM batch_jobs b
        JOIN papers p ON p.paper_id = b.paper_id
        LEFT JOIN final_counts f ON f.paper_id = b.paper_id
        ORDER BY b.created_at ASC, b.paper_id ASC
        """
    ).fetchall()


def _load_visual_stats(conn: sqlite3.Connection) -> dict[str, dict[str, int]]:
    result: dict[str, dict[str, int]] = defaultdict(_empty_visual_stats)
    queries = {
        "visual_block_count": "SELECT paper_id, COUNT(*) n FROM document_visual_blocks GROUP BY paper_id",
        "triage_count": "SELECT paper_id, COUNT(*) n FROM figure_triage_results GROUP BY paper_id",
        "crop_count": "SELECT paper_id, COUNT(*) n FROM molecule_crops GROUP BY paper_id",
        "validation_count": "SELECT paper_id, COUNT(*) n FROM molecule_crop_validations GROUP BY paper_id",
    }
    for key, sql in queries.items():
        for row in conn.execute(sql):
            result[str(row["paper_id"])][key] = int(row["n"])
    for row in conn.execute(
        """
        SELECT paper_id, SUM(CASE WHEN should_run_decimer_segmentation = 1 THEN 1 ELSE 0 END) n
        FROM figure_triage_results
        GROUP BY paper_id
        """
    ):
        result[str(row["paper_id"])]["should_segment_count"] = int(row["n"] or 0)
    return dict(result)


def _empty_visual_stats() -> dict[str, int]:
    return {
        "visual_block_count": 0,
        "triage_count": 0,
        "should_segment_count": 0,
        "crop_count": 0,
        "validation_count": 0,
    }


def _load_binding_stats(
    conn: sqlite3.Connection,
) -> dict[tuple[str, str], dict[str, int]]:
    result: dict[tuple[str, str], dict[str, int]] = defaultdict(
        lambda: {"confirmed": 0, "pending": 0}
    )
    rows = conn.execute(
        """
        SELECT paper_id, review_status, model_decision,
               reviewed_paper_material_id, model_proposed_paper_material_id
        FROM molecule_label_bindings
        """
    ).fetchall()
    for row in rows:
        review_status = str(row["review_status"] or "")
        material_id: str | None = None
        category: str | None = None
        if review_status in {"confirmed", "corrected"}:
            material_id = row["reviewed_paper_material_id"]
            category = "confirmed"
        elif review_status == "pending_review" and row["model_decision"] == "matched":
            material_id = row["model_proposed_paper_material_id"]
            category = "pending"
        if material_id and category:
            result[(str(row["paper_id"]), str(material_id))][category] += 1
    return dict(result)


def _material_status(
    *,
    material: Any,
    link: Any,
    linked_global: Any,
    task: Any,
    candidates: list[Any],
    judgments_by_candidate: dict[str, Any],
    evidence_runs: list[Any],
    evidence_items: list[Any],
    binding_stats: dict[str, int],
    visual: dict[str, int],
    name_suggestion_count: int,
) -> dict[str, str]:
    active = [candidate for candidate in candidates if candidate.status != "rejected"]
    rejected_count = len(candidates) - len(active)
    judged = [
        (candidate, judgments_by_candidate.get(candidate.structure_candidate_id))
        for candidate in active
        if judgments_by_candidate.get(candidate.structure_candidate_id) is not None
    ]
    confirmed_evidence = any(item.review_status == "confirmed" for item in evidence_items)
    ready = next(
        (
            (candidate, judgment)
            for candidate, judgment in judged
            if _candidate_has_smiles(candidate)
            and judgment.verdict in {"exact_match", "likely_match"}
            and judgment.recommended_action == "ready_for_human_accept"
        ),
        None,
    )
    blocking = next(
        (
            (candidate, judgment)
            for candidate, judgment in judged
            if judgment.verdict in {"conflict", "rejected"}
            or judgment.recommended_action == "reject_candidate"
        ),
        None,
    )
    task_requests_visual = bool(
        task
        and (
            task.next_action == "run_visual_ocsr"
            or task.assigned_strategy
            in {
                "visual_ocsr",
                "public_database_not_found",
                "public_candidate_rejected_continue_resolution",
            }
        )
    )
    task_requires_manual = bool(
        task
        and (
            task.next_action == "manual_structure_input"
            or task.assigned_strategy == "manual_structure_required"
        )
    )
    evidence_requests_figure = any(
        item.recommended_next_action == "run_figure_pipeline" for item in evidence_runs
    )
    figure_action = _figure_pipeline_action(visual)

    if any(
        candidate.status == "accepted" and _candidate_has_smiles(candidate) for candidate in active
    ):
        return _status("Ready", "None", "An accepted structure candidate is available.")
    if link and link.global_material_id and _global_has_smiles(linked_global):
        return _status("Ready", "None", "The material is linked to a structured global record.")
    if ready:
        return _status(
            "Candidate ready",
            "Accept the ready candidate",
            "A judged public structure candidate is ready for human acceptance.",
        )
    if blocking:
        return _status(
            "Identity conflict",
            "Reject the wrong candidate or add evidence",
            "The latest completed identity judgment blocks acceptance.",
        )

    scope = _structure_scope(link, task)
    if scope.get("category") == "out_of_scope_structure":
        return _status(
            "Structure not needed",
            "None",
            str(scope.get("reason") or "Outside first-phase molecular structure scope."),
        )
    if scope.get("category") == "identity_only":
        return _status(
            "Identity only",
            "None",
            str(scope.get("reason") or "Only an identity record is needed in the current scope."),
        )

    active_ocsr = [candidate for candidate in active if candidate.provider == "decimer_ocsr"]
    if active_ocsr:
        needs_correction = any(candidate.status == "needs_correction" for candidate in active_ocsr)
        return _status(
            "OCSR correction needed" if needs_correction else "OCSR candidate ready",
            "Edit the OCSR structure" if needs_correction else "Review the OCSR candidate",
            "A DECIMER candidate needs human comparison with its source crop.",
        )
    if name_suggestion_count:
        return _status(
            "Name review",
            "Review the suggested material name",
            f"{name_suggestion_count} material-name suggestion(s) are available.",
        )

    if task_requires_manual and not active:
        return _status(
            "Manual input required",
            "Enter a reviewed structure manually",
            "Public lookup and the available 2D-figure pipeline produced no reliable structure candidate.",
        )

    needs_evidence = any(
        judgment.verdict in {"ambiguous", "insufficient_evidence"}
        or judgment.recommended_action == "search_more_evidence"
        for _, judgment in judged
    ) or (bool(evidence_runs) and not confirmed_evidence)
    if needs_evidence:
        return _status(
            "Evidence ready" if confirmed_evidence else "Needs evidence",
            "Judge identity again" if confirmed_evidence else "Find or confirm identity evidence",
            "Identity evidence is incomplete or needs a new judgment.",
        )
    if active:
        if not any(
            candidate.structure_candidate_id in judgments_by_candidate for candidate in active
        ):
            return _status(
                "Needs judgment",
                "Run Identity Judge",
                f"{len(active)} active structure candidate(s) have not been judged.",
            )
        return _status(
            "Needs review",
            "Accept, reject, or edit a candidate",
            f"{len(active)} judged candidate(s) are waiting for review.",
        )

    if rejected_count:
        if (evidence_requests_figure or task_requests_visual) and figure_action:
            return _status(
                "Wrong candidate rejected",
                figure_action,
                "The wrong public candidate was rejected; continue the figure pipeline.",
            )
        if task_requests_visual:
            return _status(
                "Manual input required",
                "Enter a reviewed structure manually",
                "No usable public or 2D figure structure remains after rejection.",
            )
        return _status(
            "Wrong candidate rejected",
            "Resolve public sources again",
            "The previous candidate was rejected and no replacement is active.",
        )

    if binding_stats["confirmed"]:
        return _status(
            "OCSR ready",
            "Run OCSR",
            f"{binding_stats['confirmed']} confirmed crop binding(s) can run OCSR.",
        )
    if binding_stats["pending"]:
        return _status(
            "Needs label review",
            "Review crop-to-material binding",
            f"{binding_stats['pending']} proposed crop binding(s) need review.",
        )
    if (
        _only_paper_specific_aliases(material, link)
        and not evidence_runs
        and not task_requests_visual
    ):
        return _status(
            "Manual input required",
            "Enter a reviewed structure manually",
            "Only a paper-local label is available and no structure candidate exists.",
        )
    if link is None and task is None:
        return _status(
            "Not resolved",
            "Resolve local aliases",
            "No material link or resolution task exists.",
        )
    if task and task.status != "completed" and not task_requests_visual:
        return _status(
            "Needs lookup",
            "Resolve public sources",
            f"{task.priority} task; strategy={task.assigned_strategy}.",
        )
    if not visual["visual_block_count"]:
        return _status("Needs figures", "Run Agent Prep", "No MinerU visual blocks exist.")
    if not visual["triage_count"]:
        return _status("Needs triage", "Triage figures", "Figures have not been triaged.")
    if visual["should_segment_count"] and not visual["crop_count"]:
        return _status(
            "Needs crops", "Segment molecule crops", "Molecular figures need segmentation."
        )
    if visual["crop_count"] and not visual["validation_count"]:
        return _status(
            "Needs crop validation",
            "Validate crops",
            "Molecule crops have not been validated.",
        )
    if visual["crop_count"]:
        return _status("Needs binding", "Bind labels", "Validated crops need material binding.")
    if task_requests_visual:
        return _status(
            "Manual input required",
            "Enter a reviewed structure manually",
            "The completed figure pipeline found no usable 2D structure.",
        )
    return _status(
        "Unresolved",
        "Resolve public sources or inspect figures",
        "No accepted structure, public candidate, or usable figure crop is available.",
    )


def _status(label: str, next_action: str, detail: str) -> dict[str, str]:
    return {"label": label, "next_action": next_action, "detail": detail}


def _figure_pipeline_action(visual: dict[str, int]) -> str | None:
    if not visual["visual_block_count"]:
        return "Run Agent Prep"
    if not visual["triage_count"]:
        return "Triage figures"
    if visual["should_segment_count"] and not visual["crop_count"]:
        return "Segment molecule crops"
    if visual["crop_count"] and not visual["validation_count"]:
        return "Validate crops"
    if visual["crop_count"]:
        return "Bind labels"
    return None


def _candidate_has_smiles(candidate: Any) -> bool:
    return bool(candidate.canonical_smiles or candidate.isomeric_smiles or candidate.raw_smiles)


def _global_has_smiles(material: Any) -> bool:
    return bool(
        material and (material.canonical_smiles or material.isomeric_smiles or material.raw_smiles)
    )


def _structure_scope(link: Any, task: Any) -> dict[str, Any]:
    if link and isinstance(link.evidence.get("structure_scope"), dict):
        return link.evidence["structure_scope"]
    if task and isinstance(task.material_context.get("structure_scope"), dict):
        return task.material_context["structure_scope"]
    return {"category": "core_structure_required"}


def _only_paper_specific_aliases(material: Any, link: Any) -> bool:
    source_candidates = link.evidence.get("source_candidates", []) if link and link.evidence else []
    if source_candidates:
        return all(
            isinstance(candidate, dict) and candidate.get("is_paper_specific") is True
            for candidate in source_candidates
        )
    has_public_name = bool(
        material.full_name_in_paper or material.normalized_name or material.canonical_name
    )
    return not has_public_name and bool(
        material.abbreviation or material.paper_specific_label or material.mention_list
    )


def _material_row(
    *,
    job: sqlite3.Row,
    batch_number: int,
    batch_position: int,
    candidate_run_id: str,
    material: Any,
    status: dict[str, str],
    link: Any,
    task: Any,
    candidates: list[Any],
    judgments_by_candidate: dict[str, Any],
    evidence_runs: list[Any],
    evidence_items: list[Any],
    binding_stats: dict[str, int],
    visual: dict[str, int],
    generated_at: str,
) -> dict[str, Any]:
    active = [candidate for candidate in candidates if candidate.status != "rejected"]
    rejected = [candidate for candidate in candidates if candidate.status == "rejected"]
    judgments = [
        judgments_by_candidate[candidate.structure_candidate_id]
        for candidate in active
        if candidate.structure_candidate_id in judgments_by_candidate
    ]
    usages = material.used_in
    primary_name = next(
        (
            value
            for value in [
                material.canonical_name,
                material.full_name_in_paper,
                material.normalized_name,
                material.entity_label,
                material.abbreviation,
                material.paper_specific_label,
                *(material.mention_list or []),
            ]
            if value
        ),
        "",
    )
    return {
        "Batch Number": batch_number,
        "Batch ID": f"batch-{batch_number:04d}",
        "Batch Position": batch_position,
        "Paper ID": job["paper_id"],
        "DOI": job["doi"],
        "DOI URL": f"https://doi.org/{job['doi']}",
        "Title": job["title"] or "",
        "Journal": job["journal"] or "",
        "Publisher": job["publisher"] or "",
        "Year": job["year"] or "",
        "Paper Review Status": job["review_status"],
        "Paper Review Reason": job["review_reason"] or "",
        "Parse Status": job["parse_status"],
        "Mining Status": job["mining_status"],
        "Batch Job Status": job["job_status"],
        "Batch Job Stage": job["job_current_stage"],
        "Material ID": material.paper_material_id,
        "Primary Material Name": primary_name,
        "Entity Label": material.entity_label or "",
        "Mention List": "; ".join(material.mention_list),
        "Full Name in Paper": material.full_name_in_paper or "",
        "Normalized Name": material.normalized_name or "",
        "Canonical Name": material.canonical_name or "",
        "Abbreviation": material.abbreviation or "",
        "Paper-specific Label": material.paper_specific_label or "",
        "Material Class": material.material_class or "unknown",
        "Material Status": status["label"],
        "Recommended Next Action": status["next_action"],
        "Status Detail": status["detail"],
        "Used in Devices": "; ".join(_unique(item.device_label for item in usages)),
        "Layer Roles": "; ".join(_unique(item.layer_role for item in usages)),
        "Component Roles": "; ".join(_unique(item.component_role for item in usages)),
        "Usage Summary": "; ".join(_usage_summary(item) for item in usages),
        "Link Status": link.match_status if link else "not_resolved",
        "Link Method": link.match_method if link else "",
        "Global Material ID": link.global_material_id if link else "",
        "Task Status": task.status if task else "missing",
        "Task Current Stage": task.current_stage if task else "",
        "Task Strategy": task.assigned_strategy if task else "",
        "Task Next Action": task.next_action if task else "",
        "Task Priority": task.priority if task else "",
        "Task Error": task.error_message if task and task.error_message else "",
        "Active Candidate Count": len(active),
        "Active Candidate Providers": "; ".join(_unique(item.provider for item in active)),
        "Active Candidate Summary": "; ".join(_candidate_summary(item) for item in active),
        "Rejected Candidate Count": len(rejected),
        "Rejected Candidate Summary": "; ".join(_candidate_summary(item) for item in rejected),
        "Identity Verdicts": "; ".join(_unique(item.verdict for item in judgments)),
        "Identity Confidence": "; ".join(
            str(item.confidence) for item in judgments if item.confidence is not None
        ),
        "Identity Recommended Actions": "; ".join(
            _unique(item.recommended_action for item in judgments)
        ),
        "Identity Conflicts": "; ".join(
            _unique(conflict for item in judgments for conflict in item.conflicts)
        ),
        "Evidence Run Count": len(evidence_runs),
        "Confirmed Evidence Count": sum(
            item.review_status == "confirmed" for item in evidence_items
        ),
        "Confirmed Crop Bindings": binding_stats["confirmed"],
        "Pending Crop Bindings": binding_stats["pending"],
        "Visual Block Count": visual["visual_block_count"],
        "Figure Triage Count": visual["triage_count"],
        "Molecule Crop Count": visual["crop_count"],
        "Crop Validation Count": visual["validation_count"],
        "Candidate Run ID": candidate_run_id,
        "Generated At": generated_at,
    }


def _candidate_summary(candidate: Any) -> str:
    name = candidate.canonical_name or candidate.query_text or candidate.source_identifier or ""
    confidence = "" if candidate.confidence is None else f"{candidate.confidence:.2f}"
    return (
        f"{candidate.provider}/{candidate.resolver_name} | {name} | "
        f"status={candidate.status} | confidence={confidence} | "
        f"InChIKey={candidate.inchi_key or ''}"
    )


def _usage_summary(usage: Any) -> str:
    layer = " ".join(
        part
        for part in [
            f"layer {usage.layer_index}" if usage.layer_index is not None else "",
            usage.layer_role or "",
            usage.layer_name or "",
        ]
        if part
    )
    return " | ".join(
        part
        for part in [
            usage.device_label or "",
            layer,
            usage.component_role or "",
            usage.material_mention or "",
        ]
        if part
    )


def _unique(values: Any) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    for value in values:
        if not value:
            continue
        text = str(value)
        if text in seen:
            continue
        seen.add(text)
        result.append(text)
    return result


def _write_workbook(
    output: Path,
    *,
    rows: list[dict[str, Any]],
    summary: dict[str, Any],
    batch_stats: dict[int, Counter[str]],
) -> None:
    workbook = Workbook()
    readme = workbook.active
    readme.title = "README"
    materials = workbook.create_sheet("Unconfirmed Materials")
    statuses = workbook.create_sheet("Status Summary")
    batches = workbook.create_sheet("Batch Summary")

    readme_rows = [
        ("Report", "Unconfirmed paper materials"),
        ("Generated at", summary["generated_at"]),
        ("Database", summary["database"]),
        (
            "Batch range",
            f"{summary['criteria']['batch_range']['start']} - "
            f"{summary['criteria']['batch_range']['end'] or 'latest'}",
        ),
        ("Exported materials", summary["exported_row_count"]),
        (
            "Paper criterion",
            summary["criteria"]["paper"],
        ),
        ("Material criterion", summary["criteria"]["material"]),
        (
            "OCSR exclusions",
            ", ".join(summary["criteria"]["excluded_ocsr"]),
        ),
        (
            "OCSR materials excluded",
            summary["counters"].get("ocsr_candidate_materials_excluded", 0),
        ),
        (
            "Papers without completed material run",
            summary["counters"].get("papers_without_completed_material_run", 0),
        ),
    ]
    for row in readme_rows:
        readme.append(row)
    _style_key_value_sheet(readme)

    headers = list(rows[0]) if rows else ["No matching materials"]
    materials.append(headers)
    for item in rows:
        materials.append([_excel_value(item[header]) for header in headers])
    _style_data_sheet(materials, table_name="UnconfirmedMaterialsTable")
    if "DOI" in headers:
        doi_column = headers.index("DOI") + 1
        doi_url_column = headers.index("DOI URL") + 1
        for row_index in range(2, materials.max_row + 1):
            doi_cell = materials.cell(row=row_index, column=doi_column)
            doi_cell.hyperlink = materials.cell(row=row_index, column=doi_url_column).value
            doi_cell.style = "Hyperlink"

    statuses.append(["Material Status", "Count"])
    for label, count in summary["status_counts"].items():
        statuses.append([label, count])
    statuses.append(
        [
            "OCSR candidate ready/correction needed (excluded)",
            summary["counters"].get("ocsr_candidate_materials_excluded", 0),
        ]
    )
    _style_data_sheet(statuses, table_name="MaterialStatusSummaryTable")

    batch_headers = [
        "Batch Number",
        "Batch ID",
        "Unconfirmed Papers",
        "Papers With Materials",
        "Papers Without Materials",
        "Device-used Materials",
        "Completed Materials Excluded",
        "OCSR Candidates Excluded",
        "Exported Materials",
    ]
    batches.append(batch_headers)
    for batch_number in sorted(batch_stats):
        item = batch_stats[batch_number]
        batches.append(
            [
                batch_number,
                f"batch-{batch_number:04d}",
                item["unconfirmed_papers"],
                item["papers_with_materials"],
                item["papers_without_materials"],
                item["device_used_materials"],
                item["completed_materials_excluded"],
                item["ocsr_candidate_materials_excluded"],
                item["exported_materials"],
            ]
        )
    _style_data_sheet(batches, table_name="BatchSummaryTable")
    workbook.save(output)


def _style_key_value_sheet(sheet: Any) -> None:
    sheet.freeze_panes = "A2"
    sheet.column_dimensions["A"].width = 38
    sheet.column_dimensions["B"].width = 110
    for row in sheet.iter_rows():
        row[0].font = Font(bold=True, color="FFFFFF")
        row[0].fill = PatternFill("solid", fgColor="1F4E78")
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _style_data_sheet(sheet: Any, *, table_name: str) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    if sheet.max_row >= 2 and sheet.max_column >= 1:
        table = Table(displayName=table_name, ref=sheet.dimensions)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)
    width_overrides = {
        "Title": 55,
        "Journal": 28,
        "Publisher": 24,
        "Primary Material Name": 34,
        "Mention List": 34,
        "Full Name in Paper": 55,
        "Canonical Name": 45,
        "Status Detail": 55,
        "Usage Summary": 55,
        "Active Candidate Summary": 70,
        "Rejected Candidate Summary": 70,
        "Identity Conflicts": 70,
        "Task Error": 45,
    }
    for index, cell in enumerate(sheet[1], start=1):
        title = str(cell.value or "")
        width = width_overrides.get(title, min(max(len(title) + 3, 12), 28))
        sheet.column_dimensions[cell.column_letter].width = width


def _excel_value(value: Any) -> Any:
    if isinstance(value, str):
        value = ILLEGAL_EXCEL_XML_CHARACTERS.sub(" ", value)
        if value.startswith(("=", "+", "-", "@")):
            value = "'" + value
        if len(value) > 32_000:
            return value[:31_980] + " ...[truncated]"
    return value


if __name__ == "__main__":
    main()
